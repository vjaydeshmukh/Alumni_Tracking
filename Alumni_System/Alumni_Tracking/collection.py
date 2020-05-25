import os
import random
import re
from smtplib import SMTP
from openpyxl import load_workbook, Workbook
from django.http import HttpResponse
import datetime
from .models import *
from django.conf import settings


def list_college():
    colleges = college.objects.all()
    college_list = []
    count = 1
    for col in colleges:
        college_list.append(tuple(list([str(count), str(col.college_name)])))
        count += 1
    return tuple(college_list)


def generate_password(token=None):
    if token is None:
        get = 'abcdefghijklmnopqrstuvwxyz1234567890!@#&'
    else:
        get = 'abcdefghijklmnopqrstuvwxyz1234567890'
    password = ''
    for i in range(12):
        password += get[random.randint(0, len(get) - 1)]
    return str(password)


def sendmail(email, otp):
    connect = SMTP('smtp.gmail.com', 587)
    connect.ehlo()
    connect.starttls()
    connect.login(str('rrohitanand@gmail.com'), str(settings.MAIL_KEY))
    content = 'Subject: ' + str('OTP for login portal') + '\n\n' + str('Your six digit OTP is ') + str(
        otp) + '\n\n' + 'Regards\nAlumni\'s Call'
    connect.sendmail(str('rrohitanand@gmail.com'), str(email), content)
    connect.quit()


def generate_otp(email):
    otp = random.randrange(111111, 999999)
    sendmail(email, otp)
    return otp


def recover_mail(email, password):
    connect = SMTP('smtp.gmail.com', 587)
    connect.ehlo()
    connect.starttls()
    connect.login(str('rrohitanand@gmail.com'), str(settings.MAIL_KEY))
    content = 'Subject: ' + str('Change Password') + '\n\n' + str('Your Temporary password is ') + str(
        password) + '\n\n' + str('Please login and change your temporary password') + '\n\n' + 'Regards\nAlumni\'s Call'
    connect.sendmail(str('rrohitanand@gmail.com'), str(email), content)
    connect.quit()


def token_mail(email, token_num):
    connect = SMTP('smtp.gmail.com', 587)
    connect.ehlo()
    connect.starttls()
    connect.login(str('rrohitanand@gmail.com'), str(settings.MAIL_KEY))
    content = 'Subject: ' + str('Token Number') + '\n\n' + str(
        'Your token has been generated. Your token number is ') + str(
        token_num) + '\n\n' + 'Regards\nAlumni\'s Call'
    connect.sendmail(str('rrohitanand@gmail.com'), str(email), content)
    connect.quit()


def object_collector(request):
    try:
        profile = alumni.objects.get(user=request.user)
    except alumni.DoesNotExist:
        try:
            profile = student.objects.get(user=request.user)
        except student.DoesNotExist:
            request.session.flush()
    return profile


def post_collector(request):
    return blog.objects.filter(author__college_id=object_collector(request).college_id).order_by('-views', '-like')[:5]


def internships_collector(request):
    return internships.objects.filter(author__college_id=object_collector(request).college_id).order_by('-added')[:5]


def event_collector(request):
    return Event.objects.filter(college_id__exact=object_collector(request).college_id).order_by('-event_on')[:5]


def get_media_path(url):
    return os.path.join(os.path.join(os.getcwd(), os.path.join('media', 'resume', str(
        os.path.split(os.path.split(os.path.split(url)[0])[0])[1]), 'carrier')))


def mobile_check(request):
    if re.search('(Android)', request.META['HTTP_USER_AGENT']):
        return True


def generate_token(user, type_form):
    if os.path.exists(os.path.join(os.getcwd(), 'generated_token.xlsx')):
        wk = load_workbook(os.path.join(os.getcwd(), 'generated_token.xlsx'))
    else:
        wk = Workbook()
    token_number = generate_password(True)
    query = True
    ws = wk.active
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=2).value == user.user.email and ws.cell(row=row, column=6).value == type_form:
            token_number = ws.cell(row=row, column=5).value
            query = False
            break
        if ws.cell(row=row, column=5).value == token_number:
            token_number = generate_password(True)
    if query:
        ws.append([user.user.get_full_name(), user.user.email, user.graduate, user.college.college_name, token_number,
                   type_form, datetime.datetime.now()])
    token_mail(user.user.email, token_number)
    wk.save(os.path.join(os.getcwd(), 'generated_token.xlsx'))

    return token_number


def collect_notifications(request):
    try:
        college_id = alumni.objects.get(user=request.user).college_id
    except alumni.DoesNotExist:
        college_id = student.objects.get(user=request.user).college_id
    notify = blog.objects.filter(author__college_id=college_id).order_by('-publish_date')[:6]
    return notify
